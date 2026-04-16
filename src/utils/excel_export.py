import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


def export_to_excel(sessions, utterances_by_session: dict, output_path: str = None) -> str:
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(os.path.expanduser("~"), "Desktop", f"visuarium_{timestamp}.xlsx")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Sessions summary ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Sessions"
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    s_headers = ["세션 ID", "참여자 이름", "생성 시각", "발화 횟수", "총 발화 시간(초)"]
    for col, h in enumerate(s_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row_idx, s in enumerate(sessions, 2):
        values = [
            s["session_id"],
            s["participant_name"] or "",
            s["created_at"],
            s["utterance_count"],
            round(s["total_duration"], 2),
        ]
        for col, v in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col, value=v)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left")

    for col in range(1, len(s_headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].auto_size = True
        ws1.column_dimensions[get_column_letter(col)].width = 20

    # ── Sheet 2: All utterances ────────────────────────────────────────────
    ws2 = wb.create_sheet("Utterances")
    u_headers = ["세션 ID", "참여자 이름", "발화 순서", "발화 원문", "생성된 프롬프트",
                 "프롬프트 단어수", "발화 시간(초)", "발화 시각"]
    for col, h in enumerate(u_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    row_idx = 2
    for s in sessions:
        sid = s["session_id"]
        utterances = utterances_by_session.get(sid, [])
        for u in utterances:
            values = [
                sid,
                s["participant_name"] or "",
                u["utterance_index"],
                u["utterance_text"],
                u["prompt_text"],
                u["prompt_word_count"],
                round(u["duration_sec"], 2),
                u["spoken_at"],
            ]
            for col, v in enumerate(values, 1):
                cell = ws2.cell(row=row_idx, column=col, value=v)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="left" if col in (4, 5) else "center",
                    wrap_text=(col in (4, 5))
                )
            row_idx += 1

    col_widths = [10, 15, 10, 30, 50, 12, 14, 22]
    for col, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    wb.save(output_path)
    return output_path
